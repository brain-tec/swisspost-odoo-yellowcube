# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) 2004-2010 Tiny SPRL (<http://tiny.be>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

import base64
import tempfile
import os
from openerp import tools
from datetime import datetime

from openerp.osv import osv, fields
from openerp.tools.translate import _
import logging
_logger = logging.getLogger(__name__)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill
except:
    _logger.warning("Please install: pip install openpyxl")

from openerp.addons.bt_tax.report.account_tax_report import tax_report


class account_vat_declaration_cust(osv.osv_memory):
    _name = 'account.vat.declaration.cust'
    _description = 'Account Vat Declaration Cust'
    _inherit = "account.common.report"
    _columns = {
        'based_on': fields.selection([('invoices', 'Invoices'),
                                      #('payments', 'Payments'),
                                      ],
                                      'Based On', required=True),
        'chart_tax_id': fields.many2one('account.tax.code', 'Chart of Tax', help='Select Charts of Taxes', required=True, domain = [('parent_id','=', False)]),
        'data': fields.binary('XLS',readonly=True),
        'export_filename': fields.char('Export XLS Filename', size=128),
    }

    def _get_tax(self, cr, uid, context=None):
        taxes = self.pool.get('account.tax.code').search(cr, uid, [('parent_id', '=', False)], limit=1)
        return taxes and taxes[0] or False

    _defaults = {
        'based_on': 'invoices',
        'chart_tax_id': _get_tax,
        'export_filename': 'tax_report.xlsx'
    }

    def create_vat(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        datas = {'ids': context.get('active_ids', [])}
        datas['model'] = 'account.tax.code'
        datas['form'] = self.read(cr, uid, ids)[0]
        datas['form']['chart_tax_id'] = datas['form']['chart_tax_id'][0]
        datas['form']['chart_account_id'] = datas['form']['chart_account_id'][0]
        datas['form']['fiscalyear_id'] = datas['form']['fiscalyear_id'][0]
        if datas['form']['period_from']:
            datas['form']['period_from'] = datas['form']['period_from'][0]
        if datas['form']['period_to']:
            datas['form']['period_to'] = datas['form']['period_to'][0]
        datas['form']['company_id'] = self.pool.get('account.tax.code').browse(cr, uid, [datas['form']['chart_tax_id']], context=context)[0].company_id.id
        print 'datas: ', datas
        return {
            'type': 'ir.actions.report.xml',
            'report_name': 'account.vat.declaration.cust',
            'datas': datas,
        }
        
    def create_vat_xls(self, cr, uid, ids, context=None):
        this = self.browse(cr, uid, ids)[0]
        if context is None:
            context = {}
        data = {'ids': context.get('active_ids', [])}
        data['model'] = 'account.tax.code'
        data['form'] = self.read(cr, uid, ids)[0]
        company_id = data['form']['company_id'][0]
        t = tax_report(cr, uid, 'test', context=context)
        t.fiscalyear = data['form'].get('fiscalyear_id', False)
        t.period_ids = []
        if data['form'].get('period_from', False) and data['form'].get('period_to', False):
            t.period_ids = self.pool.get('account.period').build_ctx_periods(cr, uid, data['form']['period_from'][0], data['form']['period_to'][0])
        t.fiscalyear = data['form'].get('fiscalyear_id', False)
        if t.fiscalyear:
            t.fiscalyear = t.fiscalyear[0]
        based_on = 'invoices'
        wb = Workbook()
        # grab the active worksheet
        ws_total = wb.active
        ws_total.title = "Summary"
        lines_total = t._get_lines(based_on, company_id, False)
        if t.fiscalyear:
            ws_total.append(["Finanzjahr ", data['form']['fiscalyear_id'][1]])
        if t.period_ids:
            ws_total.append(["Periode von ", data['form']['period_from'][1], "Period bis ", data['form']['period_to'][1]])
        ws_total.append([""])
        ws_total.append([_("STEUER BEZEICHNUNG"), _("SOLL"), _("HABEN"), _("STEUERBETRAG")])
        row = ws_total.row_dimensions[ws_total.max_row]
        row.font = Font(bold=True)
        count_line_total = 0
        for line_total in lines_total:
            count_line_total += 1
            print 'line_total: ', line_total
            name = line_total['level'].replace("."," ") + line_total['level'].replace("."," ")
            if line_total['code']:
                name += tools.ustr(line_total['code']) + ' '
            name += tools.ustr(line_total['name'])
            
            tools.ustr(line_total['level'] or '') + tools.ustr(line_total['code'] or '') + tools.ustr(line_total['name'])
            
            ws_total.append([name, line_total['debit'], line_total['credit'], line_total['tax_amount']])
            if line_total['bold'] == 1:
                row = ws_total.row_dimensions[ws_total.max_row]
                row.font = Font(bold=True)
        lines = t._get_lines(based_on, company_id, True)
        for tax_line in lines:
#             print 'tax_line: ', tax_line
            if (len(tax_line['level'])==6 or len(tax_line['level'])==8) and tax_line['type']==1 and tax_line['name']:
                ws = wb.create_sheet()
                ws.title = tax_line['code']
                detail_lines = t._get_tax_lines_new(based_on, tax_line['id'])
                ws.append([_(u'DATUM'), _(u'BELEG'), _(u'TEXT'), _(u'BETRAG'), _(u'SATZ'), _(u'KONTO'), _(u'MWST BETRAG')])
                row = ws.row_dimensions[ws.max_row]
                row.font = Font(bold=True)
                total_tax_amount_new = 0
                total_amount_new = 0
                for line in detail_lines:
                  if line['res']:
                    total_tax_amount_new += line['tax_amount_new']
                    total_amount_new += line['amount_new']
                    
                    date = datetime.strptime(line['res'].date,'%Y-%m-%d').strftime('%d.%m.%Y')
                    
                    ws.append([date, 
                               line['res'].move_id.name, 
                               line['res'].partner_id.name, 
                               line['tax_amount_new'], 
                               line['res'].tax_code_id.value_for_tax_report_column_tax,
                               line['account_codes'],
                               line['amount_new']])
                    if line['res'].currency_id:
                        ws.append(['', 
                                   '', 
                                   "Fremdwährung " + line['res'].currency_id.name, 
                                   line['tax_amount_new_currency'], 
                                   '',
                                   '',
                                   line['amount_new_currency']])
        
                ws.append(['Total - ' + tax_line['name'], 
                           '', 
                           '', 
                           total_tax_amount_new, 
                           '',
                           '',
                           total_amount_new])
                row = ws.row_dimensions[ws.max_row]
                row.font = Font(bold=True)
        
#         wb.save("/home/jool1/Downloads/TESTS/sample.xlsx")
        fullpath = tempfile.mktemp(suffix='.xlsx')
        wb.save(fullpath)
        f = open(fullpath, 'r')
        file_value = f.read()
        f.close()
        os.remove(fullpath)
        try:
            self.write(cr, uid, ids,
                       {'data': base64.encodestring(file_value)},
                       context=context)
        except:
            raise
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'account.vat.declaration.cust',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
    
account_vat_declaration_cust()

#vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
